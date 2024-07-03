from openpyxl import Workbook, load_workbook
import os
from settings import EXCEL_PATH


class ExcelFile:

    @classmethod
    def create_excel(cls):
        if os.path.exists(EXCEL_PATH):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист 1'
        ws['A1'] = 'Номер заявки'
        ws['B1'] = 'Тип обращения'
        ws['C1'] = 'Описание'
        ws['D1'] = 'Контакт'
        ws['E1'] = 'Дополнительный контакт'
        ws['F1'] = 'Файлы'
        ws['G1'] = 'ID пользователя'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 100
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        wb.save(filename=EXCEL_PATH)

    @classmethod
    def paste_in_excel(cls, data, chat_id):
        cls.create_excel()
        wb = Workbook()
        wb = load_workbook(filename=EXCEL_PATH)
        ws = wb.active
        idx = str(ws.max_row + 1)
        ws['A'+idx] = data['record_id']
        ws['B'+idx] = data['type']
        ws['C'+idx] = data['descr']
        ws['D'+idx] = data.get('phone', '-')
        ws['E'+idx] = data.get('contact', '-')
        ws['F'+idx] = str(len(data['files_list']))
        ws['G'+idx] = str(chat_id)
        wb.save(filename=EXCEL_PATH)
